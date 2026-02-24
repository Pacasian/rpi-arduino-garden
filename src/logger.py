# Reading files from the Arduino
import time
import random
from datetime import datetime
from openpyxl import load_workbook, Workbook
from pathlib import Path

EXCEL_PATH = Path("data/garden_logs.xlsx")

RECORD_SECONDS = 45
CYCLE_SECONDS = 60

# This is the simulation mode, make it false if it's a simulator
SIMULATION_MODE = True

def get_fake_data():
    ''' Simulate Arduino Data '''
    moisture= random.randint(30,60)
    temp = round(random.uniform(20,30),1)
    humidity = random.randint(35,70)
    pump = "ON" if moisture < 40 else "OFF"

    return moisture, temp, humidity, pump

def ensure_excel():
    EXCEL_PATH.parent.mkdir( exist_ok=True)

    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp","SoilMoisture", "Temperature", "Humidity", "Pump"])
        wb.save(EXCEL_PATH)

    return wb,ws

def start_logger():
    print("Logger started...")

    while True:
        cycle_start = time.time()
        record_end = cycle_start + RECORD_SECONDS
        rows=[]

        while time.time() < record_end:
            if SIMULATION_MODE:
                moisture, temp, humidity, pump = get_fake_data()
            else:
                # Once the simultion off, fill this box
                pass
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            rows.append([timestamp, moisture, temp, humidity, pump])
            time.sleep(2)

        wb, ws = ensure_excel()
        for row in rows:
            ws.append(row)
        wb.save(EXCEL_PATH)

        print(f"Wrote {len(rows)} rows to Excel.")

        elapsed = time.time() - cycle_start
        time.sleep(max(0, CYCLE_SECONDS - elapsed))

